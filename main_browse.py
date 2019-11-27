'''
EXAM APPLICATION LAUNCHER developed by Mr Steven J walden
    Nov. 2019
    SAMROIYOD, PRACHUAP KIRI KHAN, THAILAND
[See license at end of file]
temp change
'''

#!/usr/bin/env python

__author__ = 'Steven Walden'
__version__ = '1.0'

import sys
import csv
import openpyxl
import os

from Create_CSV import Ui_CreateCSVWindow

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDesktopWidget
from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import Qt

class App(QtWidgets.QWidget):
	"""docstring for App"""
	def __init__(self, parent=None):
		super(App, self).__init__(parent)
		#setup app windows and theme
		
		self.dark_theme()
		self.load_data()

		self.CreateCSVWindow = QtWidgets.QDialog()
		self.Create_CSV = Ui_CreateCSVWindow()
		#connect the buttons to the methods
		self.CreateCSVWindow.browse_for_workbook = self.browse_for_workbook
		self.CreateCSVWindow.ok_button_clicked = self.ok_button_clicked
		self.CreateCSVWindow.cancel_button_clicked = self.cancel_button_clicked
		self.CreateCSVWindow.populate_sheet_cmb = self.populate_sheet_cmb

		self.Create_CSV.setupUi(self.CreateCSVWindow)

		self.screen_location()

		#limit the number of items in the student name combo box
		self.Create_CSV.SheetcomboBox.setStyleSheet("QComboBox { combobox-popup: 0; }")
		self.Create_CSV.SheetcomboBox.setMaxVisibleItems(10)
		self.Create_CSV.FileBox.setFocus()

		self.CreateCSVWindow.show()

	def load_data(self):
		self.convert_dic = {'A':1,'B':2,'C':3,'D':4,'E':5,'F':6,'G':7,'H':8,'I':9,'J':10,'k':11,'l':12,'M':13,'N':14,'O':15,'P':16,'Q':17,'R':18,'R':19,'S':20,'T':21,'U':22,'V':23,'W':24,'X':25,'Z':26,'AA':27,'AB':28,'AC':29,'AD':30,'AE':31,'AF':32,'AG':33,'AH':34,'AI':35}

	def dark_theme(self):
		app.setStyle("Fusion")

		self.dark_palette = QPalette()

		self.dark_palette.setColor(QPalette.Window,QColor(53,53,53))
		self.dark_palette.setColor(QPalette.WindowText, Qt.white)
		self.dark_palette.setColor(QPalette.Base, QColor(25, 25, 25))
		self.dark_palette.setColor(QPalette.AlternateBase, QColor(53, 53, 53))
		self.dark_palette.setColor(QPalette.ToolTipBase, Qt.white)
		self.dark_palette.setColor(QPalette.ToolTipText, Qt.white)
		self.dark_palette.setColor(QPalette.Text, Qt.white)
		self.dark_palette.setColor(QPalette.Button, QColor(53, 53, 53))
		self.dark_palette.setColor(QPalette.ButtonText, Qt.white)
		self.dark_palette.setColor(QPalette.BrightText, Qt.red)
		self.dark_palette.setColor(QPalette.Link, QColor(42, 130, 218))
		self.dark_palette.setColor(QPalette.Highlight, QColor(42, 130, 218))
		self.dark_palette.setColor(QPalette.HighlightedText, Qt.black)

		app.setPalette(self.dark_palette)

		app.setStyleSheet("QToolTip { color: #ffffff; background-color: #2a82da; border: 1px solid white; }")

	def screen_location(self):
		ag = QDesktopWidget().availableGeometry()
		sg = QDesktopWidget().screenGeometry()

		self.widget = self.CreateCSVWindow.geometry()
		x = ag.width() / 2 - self.widget.width() / 2
		y = ag.height() / 2 - self.widget.height() / 2
		self.CreateCSVWindow.move(x, y)

	def browse_for_workbook(self):
		self.workbook_name = QtWidgets.QFileDialog.getOpenFileName(None, "Select Excel File", "", "Excel Files (*.xlsx *.xlsm)")
		self.populate_file_box()

	def populate_file_box(self):
		self.Create_CSV.FileBox.setText(self.workbook_name[0])

	def populate_sheet_cmb(self):
		try:
			self.workbook = openpyxl.load_workbook(self.workbook_name[0], data_only = True)
			self.sheet_list = self.workbook.sheetnames
			self.Create_CSV.SheetcomboBox.addItems(self.sheet_list)
		except AttributeError:
			pass

	def copyRange(self, startCol, startRow, endCol, endRow, sheet):
		#Copy range of cells as a nested list
		#Takes: start cell, end cell, and sheet you want to copy from.
	    self.rangeSelected = []
	    #Loops through selected Rows
	    for i in range(startRow,endRow + 1,1):
	        #Appends the row to a RowSelected list
	        self.rowSelected = []
	        for j in range(startCol,endCol+1,1):
	            self.rowSelected.append(sheet.cell(row = i, column = j).value)
	        #Adds the RowSelected List and nests inside the rangeSelected
	        self.rangeSelected.append(self.rowSelected)

	    return self.rangeSelected

	def convert(self, val):
		return self.convert_dic[val]

	def write_csv(self):
		self.read_list =[['Name','Nicknames','Passwords'],['Choose your name','Nickname','Password']]
		with open('Student_Details_CSV_' + self.sheet_name[-4:] + '.csv', 'w') as new_file:
			csv_writer = csv.writer(new_file, delimiter = ',')

			# csv_writer.writerow(self.read_list)

			for list_detail in self.read_list:
				csv_writer.writerow(list_detail)

			#loop to write the data
			for student_detail in self.rangeSelected:
				csv_writer.writerow(student_detail)

			os.startfile('Student_Details_CSV_' + self.sheet_name[-4:] + '.csv')

	def ok_button_clicked(self):
		try:
			self.sheet_name = self.Create_CSV.SheetcomboBox.currentText()
			self.startcol = self.convert(self.Create_CSV.StartColBox.text())
			self.startrow = int(self.Create_CSV.StartRowBox.text())
			self.endcol = self.convert(self.Create_CSV.EndColBox.text())
			self.endrow = int(self.Create_CSV.EndRowBox.text())

			if self.endcol < self.startcol:
				self.Create_CSV.EndColBox.clear()
			elif self.endrow < self.startrow:
				self.Create_CSV.EndRowBox.clear()
			else:
				self.workbook_sheet = self.workbook[self.sheet_name]
				self.copyRange(self.startcol, self.startrow, self.endcol, self.endrow, self.workbook_sheet)
				self.write_csv()
		except (KeyError, ValueError):
			pass

	def cancel_button_clicked(self):
		self.CreateCSVWindow.close()

print(sys.executable)

if __name__ == '__main__':
    # print (PY_VER)
    # print (QT_VER)
    app = QtWidgets.QApplication(sys.argv)
    main_app = App()

sys.exit(app.exec_())



# Copyright (c) 2019 Steven Walden
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.