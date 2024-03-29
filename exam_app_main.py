#! python 3
'''
EXAM APPLICATION LAUNCHER developed by Mr Steven J walden
    Nov. 2019-2020
    SAMROIYOD, PRACHUAP KIRI KHAN, THAILAND
[See License.txt file]
'''
'''
Raise error reports
	video error
video/pic expand on question label
redesign tabs to 4 labels and have expand button on each
for logging use  exc_info=1 in error string to print exception info
or use exception

Timer still running after last answer
'''

__author__ = 'Mr Steven J Walden'
__version__ = '1.7.0'

import os
import sys
import time
import logging
import datetime
import pywintypes
import subprocess
import win32net
import win32file
from win32com.shell import shell, shellcon
from random import randrange, shuffle

from win32event import CreateMutex
from win32api import GetLastError
from winerror import ERROR_ALREADY_EXISTS

import csv
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QWidget, QDesktopWidget #, QApplication, QMainWindow,
#from PyQt5.QtGui import QPalette, QColor
from PyQt5.QtCore import QT_VERSION_STR, Qt, QUrl
from PyQt5.QtMultimedia import QMediaContent, QMediaPlayer
from PyQt5.QtMultimediaWidgets import QVideoWidget
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

from App_Guis import Ui_ExamLogin, Ui_ExamQuestions, Ui_StartupWindow
import methods
from methods import change_dir as cdir

PY_VER = sys.version[:3]

class App(QtWidgets.QWidget):
	"""docstring for App"""
	def __init__(self, parent=None):
		super(App, self).__init__(parent)
		#setup logger
		self.logger = logging.getLogger(__name__)
		self.logger.setLevel(logging.DEBUG)
		self.formatter = logging.Formatter('%(asctime)s:%(name)s:%(message)s')
		self.file_handler = logging.FileHandler('Error_Log')
		#self.file_handler.setLevel(logging.ERROR)
		self.file_handler.setFormatter(self.formatter)
		self.logger.addHandler(self.file_handler)
		#Load theme
		# methods.dark_theme(app)
		# methods.theme_choice(app)
		self.screen_size = QDesktopWidget().availableGeometry()
		# self.open_startup_window()
		self.network_login()
		#setup app windows and theme
		self.load_data()
		self.open_login_window()

	def network_login(self):
		#Clear any network logins
		subprocess.call("net use * /d /y", shell=True) #calls net use delete in command shell
		# #Load network location and set new login for classes, exam files and results
		try:
			with open("D:\\New Sync\\02Office & Programing\\103Exam_Project\\LL.txt", "r", encoding="utf8") as file:
				self.login_info = [line for line in file]
		except FileNotFoundError as e:
				self.logger.error(f" Cannot load the login details file! {e}")
				self.message_boxes(msg='FileNotFoundError', msg_type=3, err=e)

		self.network_location = self.login_info[0][0:-1]
		self.user_name = self.login_info[1][0:-1]
		self.pass_word = self.login_info[2][0:-1]
		self.netlogin = {'remote': self.network_location, 'local': '', 'username': self.user_name, 'password': self.pass_word}
		#new login to exam_app account
		try:
				win32net.NetUseAdd(None, 2, self.netlogin)
		except pywintypes.error as e:
				self.logger.error(str(e))
				self.message_boxes(msg='NetworkError', msg_type=3, err=e)

	def load_data(self):
		cwd = os.getcwd()
		#load class list from csv file
		with cdir(self.network_location, self.logger):	
			try:
				with open('class_list.csv','r') as csv_file:
					csv_reader = csv.DictReader(csv_file)

					self.class_list = [line['classes'] for line in csv_reader]

			except FileNotFoundError as e:
				self.logger.error(f" Cannot load the class list file! {e}")
				os.chdir(cwd)
				self.message_boxes(msg='FileNotFoundError', msg_type=2, err=e)
		#load exam list from csv file
			try:
				with open('exam_list.csv','r') as csv_file:
					csv_reader = csv.DictReader(csv_file)

					self.exam_list = [line['exams'] for line in csv_reader]

			except FileNotFoundError as e:
				self.logger.error(f" Cannot load the exam list file! {e}")
				os.chdir(cwd)
				self.message_boxes(msg='FileNotFoundError', msg_type=2, err=e)

		self.string_convert = {'A':1,'B':2,'C':3,'D':4}
		self.student_names =  []
		self.student_info, self.exam_info = {}, {}

		self.correct_answers = 0
		self.answer = 0

		self.admin = False
		self.dark_mode = False

	def read_login_csv(self, clas):
		self.student_names.clear()
		self.student_info.clear()
		self.path = f'Student_Details_CSV_M{clas[1]}-{clas[3]}.csv'
		self.csv_reader_func(path=self.path ,csv_type=0)

	def open_startup_window(self):
		self.startup_screen = Ui_StartupWindow()
		methods.screen_location(self.startup_screen, False, self.screen_size)
		self.startup_screen.setWindowTitle(f"Exam App V {__version__}")
		self.startup_screen.show()

	def open_login_window(self):
		self.login_gui = Ui_ExamLogin()
		methods.screen_location(self.login_gui, False, self.screen_size)
		self.login_gui.setWindowTitle(f"V {__version__} App login")

		#Connect the methods
		self.login_gui.buttonBox.accepted.connect(self.login_okaybutton_clicked)
		self.login_gui.buttonBox.rejected.connect(self.login_cancelbutton_clicked)
		self.login_gui.PasswordShowButton.clicked.connect(self.password_show_button_clicked)
		self.login_gui.DarkModeButton.clicked.connect(self.login_theme_choice)
		self.login_gui.InputPassword.returnPressed.connect(self.login_okaybutton_clicked)
		self.login_gui.ClassCmb.currentIndexChanged['int'].connect(self.class_name_changed)
		self.login_gui.ExamChoiceCmb.currentIndexChanged['int'].connect(self.exam_choice_changed)
		self.login_gui.StudentNameCmb.currentIndexChanged['int'].connect(self.student_name_change)

		#populate the combo boxes
		self.login_gui.ClassCmb.addItems(self.class_list)
		#limit the number of items in the student name combo box
		self.login_gui.StudentNameCmb.setStyleSheet("QComboBox { combobox-popup: 0; }")
		self.login_gui.StudentNameCmb.setMaxVisibleItems(10)
		#start with the class combobox
		self.login_gui.ClassCmb.setFocus()

		self.login_gui.show()
		# self.startup_screen.close()

	def login_theme_choice(self):
		if self.login_gui.DarkModeButton.isChecked():
			methods.dark_theme(app)
			self.login_gui.DarkModeButton.setText("Light")
			self.login_gui.Logolabel.setPixmap(QtGui.QPixmap("img/School logo75x97_grad.png"))
			self.login_gui.PasswordShowButton.setIcon(QtGui.QIcon("img/Password_Icon_20x20.png"))
			self.dark_mode = True
		else:
			methods.light_theme(app)
			self.login_gui.DarkModeButton.setText("Dark")
			self.login_gui.Logolabel.setPixmap(QtGui.QPixmap("img/School logo75x97_light.png"))
			self.login_gui.PasswordShowButton.setIcon(QtGui.QIcon("img/Password_Icon_20x20_light.png"))
			self.dark_mode = False

	def login_okaybutton_clicked(self):
		try:
			#open main window and pass variables
			self.password_input = self.login_gui.InputPassword.text()
			if self.student_number !=0 and self.exam_number !=0:
				if self.password_input == self.student_info[self.student_number]["student_password"] or self.password_input == self.student_info[0]["student_password"]:
					#Create exam results folder in the network location
					try:
						with cdir(self.network_location, self.logger):
							os.mkdir(f"M{self.year_chosen[1]}-{self.year_chosen[3]}_{self.exam_name}_results")
					except FileExistsError:
						pass
	
					#try and except check if exam taken already
					with cdir(f"{self.network_location}\\M{self.year_chosen[1]}-{self.year_chosen[3]}_{self.exam_name}_results", self.logger):
						try:
							#check to see if file has been created already if so op msg-box and deny access
							with open(f'M{self.year_chosen[1]}-{self.year_chosen[3]}_Student_{self.student_number}_{self.student_info[self.student_number]["student_nickname"]}.txt', "r") as file_object:
								self.message_boxes(msg='Exam completed already.', msg_type=4, err=None)
								#Popup msg box to save already taken exam before

						except FileNotFoundError as e:
							#If exam not taken then allow access to main window
							#open the main window
							self.login_gui.close()
							self.open_exam_window()
				else:
					self.login_gui.InputPassword.clear()
			else:
				self.login_gui.InputPassword.clear()
		except (AttributeError, KeyError):
			pass

	def password_show_button_clicked(self):
		if self.login_gui.PasswordShowButton.isDown():
			self.login_gui.InputPassword.setEchoMode(QtWidgets.QLineEdit.Normal)
		else:
			self.login_gui.InputPassword.setEchoMode(QtWidgets.QLineEdit.Password)

	def exam_choice_changed(self, exa):
		#Get exam number and name from combo box
		self.exam_number = exa
		self.exam_name = self.login_gui.ExamChoiceCmb.itemText(exa)

	def class_name_changed(self, cls):
		#receives an integer (cls) from the combo box when it changes
		self.class_name = self.login_gui.ClassCmb.itemText(cls)
		self.year_chosen = self.class_list[cls]

		if cls > 0:
			self.read_login_csv(self.class_list[cls])
			self.class_name = self.class_list[cls]
			self.login_gui.ClassLabel.setText(self.class_list[cls])
			self.login_gui.ExamChoiceCmb.clear()
			self.login_gui.ExamChoiceCmb.addItems(self.exam_list)
			self.login_gui.StudentNameCmb.clear()
			self.student_names = [self.student_info[name]["student_name"] for name in range(len(self.student_info))]
			self.login_gui.StudentNameCmb.addItems(self.student_names)
		else:
			self.login_gui.ClassLabel.setText('Class')
			self.login_gui.ExamChoiceCmb.clear()
			self.login_gui.StudentNameCmb.clear()
			self.login_gui.StudentNumber.setText('')
			self.login_gui.StudentNickname.setText('')

	def student_name_change(self, st):
		cwd = os.getcwd()
		#links student cmb box with the photo display
		self.student_number = st
		self.photo_location = (f'{self.network_location}/M{self.year_chosen[1]}-{self.year_chosen[3]}')
		with cdir(self.photo_location, self.logger):
			if st > 0:
				self.photo_path = f'{st}.png'
				if os.path.exists(self.photo_path):
					self.login_gui.StudentPhoto.setPixmap(QtGui.QPixmap(self.photo_path))
				else:
					os.chdir(cwd)
					self.logger.error(f"{self.student_info[st][student_nickname]}'s photo {str(st)}.png is missing in M{self.year_chosen[1]}-{self.year_chosen[3]} folder")
					self.login_gui.StudentPhoto.setPixmap(QtGui.QPixmap('img/blank_girl.png'))

				self.login_gui.StudentNumber.setText(str(st))
				self.login_gui.StudentNickname.setText(self.student_info[st]["student_nickname"])
			else:
				os.chdir(cwd)
				self.login_gui.StudentPhoto.setPixmap(QtGui.QPixmap('img/blank_girl.png'))
				self.login_gui.StudentNumber.setText('')
				self.login_gui.StudentNickname.setText('')

	def login_cancelbutton_clicked(self):
		app.exit()

	def open_exam_window(self):
		#Initialise exam window
		self.exam_gui = Ui_ExamQuestions(self.screen_size)
		methods.screen_location(self.exam_gui, True, self.screen_size)

		#Set any variables
		self.question_number = 1
		self.answer_state = False
		#hide back button in student mode
		if self.admin != True:
			self.exam_gui.BackButton.hide()

		#Set the theme
		if self.dark_mode:
			methods.dark_theme(app)
			self.exam_gui.DarkModeButton.setText("Light")
			self.exam_gui.SchoolLabel.setPixmap(QtGui.QPixmap("img/School logo75x97_grad.png"))
			self.exam_gui.DarkModeButton.setChecked(True)
		# else:
			# methods.light_theme(app)
			# self.exam_gui.DarkModeButton.setText("Dark")
			# self.exam_gui.SchoolLabel.setPixmap(QtGui.QPixmap("img/School logo75x97_light.png"))

		#connect butoons to methods
		self.exam_gui.LogoutButton.clicked.connect(self.logout_button_clicked)
		self.exam_gui.BackButton.clicked.connect(self.back_button_clicked)
		self.exam_gui.ForwardButton.clicked.connect(self.forward_button_clicked)
		self.exam_gui.DarkModeButton.clicked.connect(self.exam_window_theme_choice)
		self.exam_gui.AnswerButtonGroup.setId(self.exam_gui.AnswerACheckBox,1)
		self.exam_gui.AnswerButtonGroup.setId(self.exam_gui.AnswerBCheckBox,2)
		self.exam_gui.AnswerButtonGroup.setId(self.exam_gui.AnswerCCheckBox,3)
		self.exam_gui.AnswerButtonGroup.setId(self.exam_gui.AnswerDCheckBox,4)
		self.exam_gui.FalsecheckBox.hide()
		self.exam_gui.AnswerButtonGroup.buttonClicked[QtWidgets.QAbstractButton].connect(self.check_answer)
		self.exam_gui.mediaPlayer.stateChanged.connect(self.repeat_video)

		self.read_exam_questions_csv()
		#set allowed time from exam questions CSV
		self.allowed_time = (int(self.exam_info[0]["answer_b"]) * 600)
		#Set the times for the exam
		self.start_time = datetime.datetime.today()
		self.end_time = self.start_time + datetime.timedelta(hours=0, minutes=int(self.allowed_time / 600))

		#set the text etc
		self.exam_gui.setWindowTitle(f'{self.exam_info[0]["question"]} {self.exam_info[0]["answer_a"]} Questions')
		self.exam_gui.ExamTitle.setText(f'{self.exam_info[0]["question"]}\n{self.exam_info[0]["answer_a"]}')
		self.exam_gui.StartTime.setText(self.start_time.strftime("%H:%M:%S"))
		self.exam_gui.EndTime.setText(self.end_time.strftime("%H:%M:%S"))
		self.exam_gui.ClassLabel.setText(self.class_name)
		self.exam_gui.StudentNumberLabel.setText(str(self.student_number))
		self.exam_gui.StudentNicknameLabel.setText(self.student_info[self.student_number]["student_nickname"])
		self.exam_gui.StudentNameLabel.setText(self.student_info[self.student_number]["student_name"])
		self.exam_gui.OutOfQuestionLabel.setText(f"/{len(self.exam_info) -1}")

		with cdir(self.photo_location, self.logger):
			try:
				if os.path.exists(self.photo_path):
					self.exam_gui.StudentPhotoLabel.setPixmap(QtGui.QPixmap(self.photo_path))
				else:
					self.exam_gui.StudentPhotoLabel.setPixmap(QtGui.QPixmap('img/blank_girl.png'))
			except Exception as e:
				self.logger.error(str(e))

		#self.exam_gui.tabWidget.setCurrentIndex(0)

		self.exam_gui.TimeLeftProgressBar.setMaximum(self.allowed_time)
		self.exam_gui.TimeLeftProgressBar.setMinimum(0)
		self.exam_gui.TimeLeftProgressBar.setValue(self.allowed_time)
		self.exam_gui.MinLeftLabel.setText(f"{int(self.allowed_time / 600)} Min Left")

		#Create list of answerlabels and answer texts
		self.answer_label_list = [self.exam_gui.AnswerTextA,self.exam_gui.AnswerTextB,self.exam_gui.AnswerTextC,self.exam_gui.AnswerTextD]

		#Create a list of question numbers and shuffle them
		self.quest_seq = [q_num for q_num in range(1, len(self.exam_info))]
		shuffle(self.quest_seq)

		#Show window
		self.populate_boxes(self.quest_seq[self.question_number - 1]) #pass the random question from a list

		# self.login_gui.hide()
		self.exam_gui.show()
		self.counters()

	def exam_window_theme_choice(self):
		if self.exam_gui.DarkModeButton.isChecked():
			methods.dark_theme(app)
			self.exam_gui.DarkModeButton.setText("Light")
			self.exam_gui.SchoolLabel.setPixmap(QtGui.QPixmap("img/School logo75x97_grad.png"))
		else:
			methods.light_theme(app)
			self.exam_gui.DarkModeButton.setText("Dark")
			self.exam_gui.SchoolLabel.setPixmap(QtGui.QPixmap("img/School logo75x97_light.png"))

	def csv_reader_func(self, path, csv_type):
		'''
		Function to get student details information from a .CSV file
		Reads a .CSV file from the network directory and appends the infomation to 3 lists.
		Arguments:
			clas {[integer]} -- [This is the selected school year number to be used as the first part of the .CSV filename i.i M1
		'''
		cwd = os.getcwd()
		with cdir(self.network_location, self.logger):
			#open the correct csv file for each login/class name
			#Test to see if the file exsists if not raise error and close program
			try:
				with open(path,'r') as csv_file:
					csv_reader = csv.DictReader(csv_file)

					# for line in csv_reader:
					if csv_type == 0: #reads student details csv
						self.student_info = {int(line['Student number']): {"student_name": line['Name'], "student_nickname": line['Nickname'], "student_password": line['Password']} for line in csv_reader}
					else: #reads exam questions csv
						self.exam_info = {int(line['QuestionNumber']): {"question": line['Questions'], "answer_a": line['AnswerA'], "answer_b": line['AnswerB'], "answer_c": line['AnswerC'], "answer_d": line['AnswerD'], "correct_answer": line['Rightanswer'], "photo_question": line['Photoquestion']} for line in csv_reader}

			except FileNotFoundError as e:
				self.logger.error(f" Can not find the file {path}")
				os.chdir(cwd)
				self.message_boxes(msg='FileNotFoundError', msg_type=2, err=e)

	def read_exam_questions_csv(self):
		self.exam_info.clear()

		self.path = f'{self.exam_name}\\{self.exam_name}_Questions.csv'
		self.csv_reader_func(path=self.path ,csv_type=1)

	def counters(self):
		self.scroll_thread = methods.ScrollThread(parent=None, alloted_time=self.allowed_time)
		self.scroll_thread.start()
		self.scroll_thread.time_value.connect(self.set_progress_bar)

		self.time_thread = methods.TimeThread(parent=None, alloted_time=self.allowed_time)
		self.time_thread.start()
		self.time_thread.time_value.connect(self.set_time_label)

	def set_progress_bar(self, left_time):
		self.exam_gui.TimeLeftProgressBar.setValue(left_time)
		#exits app and saves current score
		if left_time <= 0:
			self.save_results()
			self.message_boxes(msg='Time finished!', msg_type=1, err=None)

	def set_time_label(self, left_time):
		self.exam_gui.MinLeftLabel.setText(f"{left_time} Min Left")

	def logout_button_clicked(self):
		self.message_boxes(msg='Logout?', msg_type=0, err=None)

	def message_boxes(self, msg, msg_type, err):
		self.msgbox = QMessageBox()
		self.msgbox.setWindowIcon(QtGui.QIcon("img/ep_program_logo_user_acc_zrP_icon.ico"))
		self.msgbox.setWindowTitle(msg)
		self.msgbox.setDefaultButton(QMessageBox.Ok)
		self.msgbox.setWindowFlags(QtCore.Qt.CustomizeWindowHint | QtCore.Qt.WindowTitleHint)

		if msg_type == 4:
			self.msgbox.setText("Only one attempt allowed!.")
			self.msgbox.setIcon(QMessageBox.Warning)
			self.msgbox.setStandardButtons(QMessageBox.Ok)
		if msg_type == 3:
			self.msgbox.setText(f"Please contact your teacher.\n{err}")
			self.msgbox.setIcon(QMessageBox.Critical)
			self.msgbox.setStandardButtons(QMessageBox.Ok)
		if msg_type == 2:
			self.msgbox.setText(f"Please contact your teacher.\n{err}")
			self.msgbox.setIcon(QMessageBox.Critical)
			self.msgbox.setStandardButtons(QMessageBox.Ok)
		if msg_type == 1:
			self.msgbox.setText(f'Your score is {self.correct_answers}/{len(self.exam_info)-1}')
			self.msgbox.setIcon(QMessageBox.Information)
			self.msgbox.setStandardButtons(QMessageBox.Ok)
			#clear list
			self.result_list.clear()
			self.correct_answers = 0
		if msg_type == 0:
			self.msgbox.setText('Caution! \n You will loose your score.')
			self.msgbox.setIcon(QMessageBox.Warning)
			self.msgbox.setStandardButtons(QMessageBox.Ok|QMessageBox.Cancel)
			self.correct_answers = 0

		if self.msgbox.exec() == QMessageBox.Ok:
			self.allowed_time = 20 * 600
			self.question_number = 1
			if msg_type == 3:
				sys.exit()
			if msg_type < 2:
				self.scroll_thread.is_running = False
				self.time_thread.is_running = False
				self.exam_gui.close()
				self.open_login_window()

	def save_results(self):
		self.time_finished = datetime.datetime.today()
		#Store score in a list
		self.result_list = [self.student_number, self.student_info[self.student_number]["student_name"], self.student_info[self.student_number]["student_nickname"], self.correct_answers, self.start_time.strftime("%d/%m/%Y"), self.start_time.strftime("%H:%M:%S"), self.time_finished.strftime("%H:%M:%S")]

		#Save a copy of the result to the documents folder
		doc_folder = shell.SHGetFolderPath(0, shellcon.CSIDL_PERSONAL, None, 0)
		self.results_backup_filename = f'{self.exam_info[0]["question"]} {self.exam_info[0]["answer_a"]} Student {self.student_number} results.txt'
		with cdir(doc_folder, self.logger):
			try:
				with open(self.results_backup_filename, 'w') as results_file:
					results_file.write(f'{self.student_number}-{self.student_info[self.student_number]["student_name"]}-{self.student_info[self.student_number]["student_nickname"]} Score= {self.correct_answers}')
			except Exception as e:
				self.logger.error(f" Cannot save a copy of the results to {doc_folder} because {e}")
				os.chdir(cwd)


		#Check for exsisting excel file
		self.results_filename = f'{self.exam_info[0]["question"]} {self.exam_info[0]["answer_a"]} results.xlsx'
		with cdir(f"{self.network_location}\\M{self.year_chosen[1]}-{self.year_chosen[3]}_{self.exam_name}_results", self.logger): #
			try:
				self.results_wb = load_workbook(filename = self.results_filename) #opening the file
				self.write_to_result_wb()
			except Exception as network_error:
				self.results_wb = Workbook() #create the workbook then write to workbook method and save
				self.write_to_result_wb()
				self.logger.error(f" Error: {network_error}")
				self.logger.error(f" Created: {self.results_filename} in {os.getcwd()}")

		#Save exam completed check file
		with cdir(f"{self.network_location}\\M{self.year_chosen[1]}-{self.year_chosen[3]}_{self.exam_name}_results", self.logger):
			try:
				with open(f'M{self.year_chosen[1]}-{self.year_chosen[3]}_Student_{self.student_number}_{self.student_info[self.student_number]["student_nickname"]}.txt', "w") as check_file:
					check_file.write("Exam completed")
			except Exception as e:
				self.logger.error(f" Cannot save the check file to {doc_folder} because {e}")
				os.chdir(cwd)

	def save_running_result(self):
		self.running_results_filename = f'{self.exam_info[0]["question"]}_{self.exam_info[0]["answer_a"]}_Student_{self.student_number}_{self.student_info[self.student_number]["student_name"]}_{self.student_info[self.student_number]["student_nickname"]}_running_results.txt'
		self.text_to_write = f"Question number {self.quest_seq[self.question_number - 1]} = {self.answer_state} Total score= {self.correct_answers}"
		with cdir(f"{self.network_location}\\M{self.year_chosen[1]}-{self.year_chosen[3]}_{self.exam_name}_results", self.logger):
			try:
				self.append_new_line_to_file(self.running_results_filename, self.text_to_write)
			except FileNotFoundError as e:
				self.logger.error(f" Cannot load the running result file! {e}")
				os.chdir(cwd)

	def append_new_line_to_file(self, file_name, text_to_append):
		"""Append given text as a new line at the end of file"""
		# Open the file in append & read mode ('a+')
		with open(file_name, "a+") as file_object:
		# Move read cursor to the start of file.
			file_object.seek(0)
	        # If file is not empty then append '\n'
			data = file_object.read(100)
			if len(data) > 0:
				file_object.write("\n")
	        # Append text at the end of file
			file_object.write(text_to_append)

	def write_to_result_wb(self):
		self.header_list = ['Number','Name','Nickname','Score','Day Taken','Time Started','Time Finished']
		#Write list contents to file
		self.work_sheet = self.results_wb.active
		self.work_sheet.title = self.exam_info[0]["question"]
		#self.work_sheet.column_dimensions['B'].width = 30 #Check width from length of string
		#writes the header and the results to the worksheet
		for col in range(len(self.result_list)):
			self.header_cell = self.work_sheet.cell(row=1, column=col+1, value=self.header_list[col])
			self.header_cell.font = Font(size=12, bold=True, italic=True)

			self.work_sheet.cell(row=self.student_number+1, column=col+1, value=self.result_list[col])

		try:
			self.results_wb.save(filename = self.results_filename)
		except PermissionError as e:
			self.logger.error(f" File was still open: {e}")
			#save the results to the running result file
			try:
				self.text_to_write = f'{self.student_number}-{self.student_info[self.student_number]["student_name"]}-{self.student_info[self.student_number]["student_nickname"]} Score= {self.correct_answers}'
				self.append_new_line_to_file(self.running_results_filename, self.text_to_write)
			except Exception as e:
				pass
			self.message_boxes(msg='PermissionError', msg_type=2, err=f'File is still open {e}')


		self.results_wb.close()

	def forward_button_clicked(self):
		if self.answered > 0:
			if self.answer_state:
				self.correct_answers += 1
			self.save_running_result()
			self.question_number += 1
			if self.question_number > (len(self.exam_info) -1):
				self.question_number = (len(self.exam_info) -1)
				self.save_results()
				self.message_boxes(msg='Exam finished!', msg_type=1, err=None)
			self.exam_gui.tabWidget.setCurrentIndex(0)
			self.exam_gui.FalsecheckBox.setChecked(True)

			self.populate_boxes(self.quest_seq[self.question_number - 1]) #pass the random question from a list

	def back_button_clicked(self):
		self.question_number -= 1
		if self.question_number <= 0:
			self.question_number = 1
		self.exam_gui.tabWidget.setCurrentIndex(0)
		self.populate_boxes(self.question_number)

	def populate_boxes(self, quest):
		self.exam_gui.mediaPlayer.stop()
		#Set text on exam questions and answers from list/csv
		self.answered = 0
		self.exam_gui.QuestionNumber.setText(str(self.question_number))
		self.exam_gui.Questions.setText(self.exam_info[quest]["question"])
		#Loop through label & answer dict and populate the labels with the answers
		self.exam_info_key_list = ["answer_a", "answer_b", "answer_c", "answer_d"]
		answernum = 0
		for answer_label in self.answer_label_list:
			if len(self.exam_info[quest][self.exam_info_key_list[answernum]]) > 4 and self.exam_info[quest][self.exam_info_key_list[answernum]][-4:] == '.jpg':
				with cdir(f"{self.network_location}/{self.exam_name}", self.logger):
					# myPixmap = QtGui.QPixmap(self.exam_answers_list[num][quest])
					# myScaledPixmap = myPixmap.scaled(answer_label.size(), Qt.KeepAspectRatio)
					# answer_label.setPixmap(myScaledPixmap)
					answer_label.setPixmap(QtGui.QPixmap(self.exam_info[quest][self.exam_info_key_list[answernum]]))
					answer_label.setScaledContents(True)#check to see about scaling
			else:
				answer_label.setText(self.exam_info[quest][self.exam_info_key_list[answernum]])
			answernum += 1

		#Set video media
		fileName = f'{self.network_location}/{self.exam_name}/{self.exam_info[quest]["photo_question"]}'
		try:
			if self.exam_info[quest]["photo_question"] != 'None':
				self.exam_gui.mediaPlayer.setMedia(QMediaContent(QUrl.fromLocalFile(fileName)))
				self.exam_gui.mediaPlayer.play()

		except Exception as e:
			pass

	def repeat_video(self, video_state):

		if video_state == 0 and self.exam_info[self.quest_seq[self.question_number - 1]]["photo_question"] != 'None':
			self.exam_gui.mediaPlayer.play()

	def check_answer(self, btn):
		self.answered = self.exam_gui.AnswerButtonGroup.checkedId()
		if self.answered == self.string_convert[self.exam_info[self.quest_seq[self.question_number - 1]]["correct_answer"]]:
			self.answer_state = True
		else:
			self.answer_state = False

handle = CreateMutex(None, 1, 'A unique mutex name')
print(sys.executable)

if __name__ == '__main__':
	if GetLastError() == ERROR_ALREADY_EXISTS:
		sys.exit(1) #exit if app instance already exists
	else:
		print("Qt version:", QT_VERSION_STR)
		print("Author:", __author__)
		print("App version:",__version__)
		
		app = QtWidgets.QApplication(sys.argv)
		main_app = App()

	sys.exit(app.exec_())

